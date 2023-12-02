

import openpyxl
import sendmail as sm
import yearextractor as YS




def emailextract(emailxlpath,oppath,docpath):

  workbook = openpyxl.load_workbook(emailxlpath)

  sheet = workbook.active

  total_rows = sheet.max_row
  
  ss=YS.yearsemextractor(docpath)

# Iterate over rows starting from the 2nd row
  for row_index in range(2, total_rows + 1):
    # Get the value from the first column (column index 1)
    name = sheet.cell(row=row_index, column=1).value
    print(name)
    
    email = sheet.cell(row=row_index, column=4).value
    print(email)
    
    if email is not None:
      sm.SM(name,email,oppath,ss)
      #print("------check1------")

# Close the workbook
  workbook.close()